import re
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RE_EMOJI = re.compile(
    "["
    "\U0001F600-\U0001F64F"
    "\U0001F300-\U0001F5FF"
    "\U0001F680-\U0001F6FF"
    "\U0001F1E0-\U0001F1FF"
    "\U00002702-\U000027B0"
    "\U000024C2-\U0001F251"
    "\U0001F926-\U0001FA9F"
    "\U00010000-\U0010FFFF"
    "\u2600-\u2B55"
    "\u200d"
    "\ufe0f"
    "\u3030"
    "]+",
    flags=re.UNICODE,
)

COLUMNS = ["id", "text", "likes", "dislikes", "SOURCE", "RELATION", "CATEGORY"]

def make_row(row_id, text, likes, dislikes, source, relation, category=""):
    return {
        "id": row_id,
        "text": normalize_whitespace(text),
        "likes": likes,
        "dislikes": dislikes,
        "SOURCE": source,
        "RELATION": relation,
        "CATEGORY": category,
    }

def normalize_whitespace(text):
    if not text:
        return ""
    text = RE_EMOJI.sub("", str(text))
    return re.sub(r"\s+", " ", text).strip()

def is_valid_text(text):
    return bool(text and isinstance(text, str) and len(text.strip()) >= 2)

def clean_dataframe(df):
    if df.empty:
        print("Warning: No data collected.")
        return pd.DataFrame(columns=COLUMNS)
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = "" if col == "CATEGORY" else 0
    df["text"] = df["text"].apply(lambda x: normalize_whitespace(str(x)) if pd.notna(x) else "")
    df = df[df["text"].apply(is_valid_text)].copy()
    df = df.drop_duplicates(subset=["text", "SOURCE"], keep="first").copy()
    df["likes"]    = pd.to_numeric(df["likes"],    errors="coerce").fillna(0).astype(int)
    df["dislikes"] = pd.to_numeric(df["dislikes"], errors="coerce").fillna(0).astype(int)
    before = len(df)
    df = df[(df["likes"] > 0) | (df["dislikes"] > 0)].copy()
    dropped = before - len(df)
    if dropped:
        print(f"  Filtered {dropped} comments with no reactions (0 likes, 0 dislikes)")
    df["RELATION"] = pd.to_numeric(df["RELATION"], errors="coerce").fillna(0).astype(int)
    df["CATEGORY"] = df["CATEGORY"].fillna("").astype(str)
    return df.reset_index(drop=True)[COLUMNS]

def resolve_output_path(filepath: str) -> str:
    if not os.path.exists(filepath):
        return filepath
    try:
        with open(filepath, "a"):
            pass
        return filepath
    except PermissionError:
        print(f"\n  ⚠ '{filepath}' is open in Excel. Close it and press ENTER to continue.")
        input()
        return resolve_output_path(filepath)

def save_to_excel(df: pd.DataFrame, filepath: str) -> None:
    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
    filepath = resolve_output_path(filepath)
    df.to_excel(filepath, index=False, engine="openpyxl")

    wb = load_workbook(filepath)
    ws = wb.active

    header_font  = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", start_color="2E6DA4")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BBBBBB")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        cell.border = cell_border

    light_fill = PatternFill("solid", start_color="EBF3FB")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font   = Font(name="Arial", size=10)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = light_fill

    col_widths = {"id": 6, "text": 60, "likes": 8, "dislikes": 10,
                  "SOURCE": 15, "RELATION": 10, "CATEGORY": 18}
    for i, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 20)

    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"  Saved → {filepath}")
